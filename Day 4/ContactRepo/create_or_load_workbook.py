from openpyxl import Workbook, load_workbook
def create_or_load_workbook(filename):
	try:
		wb = load_workbook(filename)
	except FileNotFoundError:
		wb = Workbook()
		sheet = wb.active
		sheet.title = "Contact Form"
		sheet.append(["id","Name", "Email", "Country Code", "Phone Number"])
		wb.save(filename)
	return wb