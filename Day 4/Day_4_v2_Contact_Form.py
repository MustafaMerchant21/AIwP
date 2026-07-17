import tkinter as tk
from tkinter import ttk, messagebox
from collections import Counter

from openpyxl import load_workbook
import ContactRepo.config as config
from ContactRepo.create_or_load_workbook import create_or_load_workbook as clw

# ==================================================================
# Theme / colors
# ==================================================================
FONT = getattr(config, "FONT_FAMILY", "Segoe UI")

BG            = "#EEF1F6"   # app bg
CARD          = "#FFFFFF"
BORDER        = "#E2E8F0"
BORDER_FOCUS  = "#6366F1"

HEADER_BG     = "#0F172A"
HEADER_FG     = "#F8FAFC"
HEADER_MUTED  = "#94A3B8"
HEADER_CHIP   = "#1E293B"

PRIMARY       = "#4F46E5"   # main accent, used sparingly
PRIMARY_DARK  = "#4338CA"
PRIMARY_TINT  = "#EEF2FF"
PRIMARY_DEEP  = "#3730A3"

TEXT          = "#0F172A"
TEXT_2        = "#475569"
MUTED         = "#94A3B8"

SUCCESS       = "#047857"
SUCCESS_TINT  = "#ECFDF5"
SUCCESS_HOVER = "#D1FAE5"

DANGER        = "#DC2626"
DANGER_DEEP   = "#B91C1C"
DANGER_TINT   = "#FEF2F2"
DANGER_HOVER  = "#FEE2E2"

NEUTRAL_TINT  = "#F1F5F9"
NEUTRAL_HOVER = "#E2E8F0"
NEUTRAL_TEXT  = "#334155"

ROW_EVEN      = "#FFFFFF"
ROW_ODD       = "#F8FAFC"
ROW_SELECTED  = "#E0E7FF"
ROW_SEL_FG    = "#1E1B4B"

HEAD_BG       = "#F8FAFC"   # table heading bg, deliberately not accent-colored
HEAD_FG       = "#475569"

SEARCH_PLACEHOLDER = "Search name, email, or phone…"


# ==================================================================
# Data helpers
# ==================================================================
def get_workbook_sheet():
    """Open (or create) the workbook, return it w/ the active sheet."""
    wb = clw(config.FILE_NAME)
    return wb, wb.active


def get_next_id(sheet):
    """Next id = max existing int id + 1, starts at 1 if sheet's empty."""
    ids = [row[0].value for row in sheet.iter_rows(min_row=2) if isinstance(row[0].value, int)]
    return (max(ids) + 1) if ids else 1


def id_exists(sheet, record_id_int, exclude_id=None):
    """Check for a clashing record id. exclude_id skips the row being edited."""
    for row in sheet.iter_rows(min_row=2):
        existing_id = row[0].value
        if not isinstance(existing_id, int):
            continue
        if existing_id == record_id_int and existing_id != exclude_id:
            return True
    return False


def email_exists(sheet, email, exclude_id=None):
    """Case-insensitive dup check on email, skips exclude_id on update."""
    email_lower = email.strip().lower()
    for row in sheet.iter_rows(min_row=2):
        existing_id = row[0].value
        existing_email = row[2].value
        if existing_email and str(existing_email).strip().lower() == email_lower:
            if existing_id != exclude_id:
                return True
    return False


def phone_number_exists(sheet, phone_number, exclude_id=None):
    """Dup check on phone number, skips exclude_id on update."""
    phone_number_str = str(phone_number).strip()
    for row in sheet.iter_rows(min_row=2):
        existing_id = row[0].value
        existing_phone = row[4].value
        if existing_phone and str(existing_phone).strip() == phone_number_str:
            if existing_id != exclude_id:
                return True
    return False


# ==================================================================
# Validation + form state
# ==================================================================
FIELD_WRAPPERS = {}   # entry widget -> its bordered wrapper frame (set in make_field)


def reset_field_borders():
    """Wipe any red error outlines back to the default border color."""
    for entry in (name_entry, email_entry):
        wrapper = FIELD_WRAPPERS.get(entry, entry)
        wrapper.config(highlightbackground=BORDER, highlightcolor=BORDER_FOCUS)
    phone_group.config(highlightbackground=BORDER, highlightcolor=BORDER_FOCUS)


def mark_invalid(entry):
    """Flag a field red + focus it. Phone group is highlighted as one unit."""
    if entry in (country_code_entry, phone_number_entry):
        phone_group.config(highlightbackground=DANGER, highlightcolor=DANGER)
    else:
        wrapper = FIELD_WRAPPERS.get(entry, entry)
        wrapper.config(highlightbackground=DANGER, highlightcolor=DANGER)
    entry.focus_set()


def validate_data(name, email, country_code, phone_number):
    """Run every field check in order, stop + flag on first failure."""
    reset_field_borders()

    name = name.strip()
    email = email.strip()
    country_code = country_code.strip()
    phone_number = phone_number.strip()

    if not name:
        mark_invalid(name_entry)
        set_status("Full name is required.", is_error=True)
        return False
    if len(name) < 3:
        mark_invalid(name_entry)
        set_status("Full name must be at least 3 characters.", is_error=True)
        return False
    if not config.NAME_PATTERN.fullmatch(name):
        mark_invalid(name_entry)
        set_status("Use letters, spaces, apostrophes, hyphens, or dots in the name.", is_error=True)
        return False
    if not email:
        mark_invalid(email_entry)
        set_status("Email is required.", is_error=True)
        return False
    if len(email) > 254:
        mark_invalid(email_entry)
        set_status("Email is too long.", is_error=True)
        return False
    if not config.EMAIL_PATTERN.fullmatch(email):
        mark_invalid(email_entry)
        set_status("Enter a valid email address.", is_error=True)
        return False
    if not phone_number:
        mark_invalid(phone_number_entry)
        set_status("Contact number is required.", is_error=True)
        return False
    if country_code and not config.COUNTRY_CODE_PATTERN.fullmatch(country_code):
        mark_invalid(country_code_entry)
        set_status("Country code must be 1 to 4 digits and cannot start with 0.", is_error=True)
        return False
    if not config.PHONE_PATTERN.fullmatch(phone_number):
        mark_invalid(phone_number_entry)
        set_status("Phone number must contain 10 to 15 digits.", is_error=True)
        return False
    return True


def validate_record_id(record_id):
    """Confirm record id is present + a valid positive int (for update/delete)."""
    record_id = record_id.strip()
    if not record_id:
        set_status("Select a record from the list or enter an ID.", is_error=True)
        return False
    if not config.RECORD_ID_PATTERN.fullmatch(record_id):
        set_status("Record ID must be a positive whole number.", is_error=True)
        return False
    return True


# ==================================================================
# Status bar (auto-clears after 6 seconds)
# ==================================================================
_status_job = None


def set_status(text, is_error=False, sticky=False):
    """Update the status pill, auto-clears after 6s unless sticky=True."""
    global _status_job
    status_label.config(
        text=("  ⚠  " if is_error else "  ✓  ") + text,
        bg=DANGER_TINT if is_error else SUCCESS_TINT,
        fg=DANGER_DEEP if is_error else SUCCESS,
    )
    status_bar.config(bg=DANGER if is_error else SUCCESS)
    if _status_job is not None:
        root.after_cancel(_status_job)   # kill prev timer so they don't stack up
        _status_job = None
    if not sticky:
        _status_job = root.after(6000, _reset_status)


def _reset_status():
    """Timer callback, puts status bar back to the default 'Ready.' state."""
    global _status_job
    _status_job = None
    status_label.config(text="  ✓  Ready.", bg=SUCCESS_TINT, fg=SUCCESS)
    status_bar.config(bg=SUCCESS)


def clear_form(event=None):
    """Reset form back to a blank 'new record' state."""
    reset_field_borders()
    record_id_var.set("")
    name_var.set("")
    email_var.set("")
    country_code_var.set("91")
    phone_number_var.set("")
    table.selection_remove(table.selection())
    mode_label.config(text=" NEW RECORD ", bg=NEUTRAL_TINT, fg=TEXT_2)
    name_entry.focus_set()


def get_form_values():
    """Grab + strip current values from all form fields as a tuple."""
    return (
        record_id_var.get().strip(),
        name_var.get().strip(),
        email_var.get().strip(),
        country_code_var.get().strip(),
        phone_number_var.get().strip(),
    )


# ==================================================================
# Table / search / sorting
# ==================================================================
cached_rows = []
_sort_state = {"col": None, "reverse": False}
COLUMNS = ("id", "name", "email", "country_code", "phone_number")
HEADINGS = {"id": "ID", "name": "NAME", "email": "EMAIL", "country_code": "CODE", "phone_number": "PHONE"}


def load_rows():
    """Read all data rows from the sheet (row 1 is header, skipped)."""
    wb, sheet = get_workbook_sheet()
    return list(sheet.iter_rows(min_row=2, values_only=True))


def _sorted(rows):
    """Sort rows by whatever col/direction is active. id sorts numeric, rest as text."""
    col = _sort_state["col"]
    if col is None:
        return rows
    idx = COLUMNS.index(col)
    if col == "id":
        # non-int/None ids just get pushed to the bottom, shouldn't happen but eh
        key = lambda r: (r[idx] is None, r[idx] if isinstance(r[idx], int) else 0)
    else:
        key = lambda r: str(r[idx] or "").lower()
    return sorted(rows, key=key, reverse=_sort_state["reverse"])


def _update_heading_arrows():
    """Stick a ▴/▾ on whichever column header is the active sort col."""
    for col in COLUMNS:
        label = HEADINGS[col]
        if col == _sort_state["col"]:
            label += "  ▾" if _sort_state["reverse"] else "  ▴"
        table.heading(col, text=label)


def sort_by(col):
    """Same col clicked again = flip dir, new col = sort asc first."""
    if _sort_state["col"] == col:
        _sort_state["reverse"] = not _sort_state["reverse"]
    else:
        _sort_state["col"] = col
        _sort_state["reverse"] = False
    _update_heading_arrows()
    apply_search_filter()


def render_table(rows):
    """Repaint the treeview w/ given rows + toggle the empty-state msg."""
    rows = _sorted(rows)
    table.delete(*table.get_children())
    for index, row in enumerate(rows):
        tag = "evenrow" if index % 2 == 0 else "oddrow"
        table.insert("", tk.END, values=row, tags=(tag,))
    count_label.config(text=f"  {len(rows)} contact{'s' if len(rows) != 1 else ''}  ")

    if rows:
        empty_label.place_forget()
    else:
        # diff msg depending on empty sheet vs a search that just found nothing
        msg = "No matches found." if _search_query() else "No contacts yet.\nAdd your first record on the left."
        empty_label.config(text=msg)
        empty_label.place(relx=0.5, rely=0.4, anchor="center")


def _search_query():
    """Current search text, ignoring the placeholder string."""
    text = search_var.get().strip()
    if text == SEARCH_PLACEHOLDER:
        return ""
    return text.lower()


def apply_search_filter(event=None):
    """Filter cached rows by name/email/phone match + re-render."""
    query = _search_query()
    if not query:
        render_table(cached_rows)
        clear_search_btn.grid_remove()
        return
    clear_search_btn.grid()
    filtered = [
        row for row in cached_rows
        if query in str(row[1]).lower() or query in str(row[2]).lower() or query in str(row[4]).lower()
    ]
    render_table(filtered)


def clear_search():
    """Reset search box to placeholder + show all rows again."""
    search_var.set("")
    _apply_placeholder()
    apply_search_filter()
    table.focus_set()


def update_stats():
    """Recompute the 'N domains · top code' caption from the full dataset."""
    domains = {str(r[2]).split("@")[-1].lower() for r in cached_rows if r[2] and "@" in str(r[2])}
    codes = [str(r[3]) for r in cached_rows if r[3]]
    if not codes:
        stats_label.config(text="")
        return
    top_code, _ = Counter(codes).most_common(1)[0]
    domain_word = "domain" if len(domains) == 1 else "domains"
    stats_label.config(text=f"{len(domains)} unique {domain_word}  ·  +{top_code} is the most used code")


def refresh_table():
    """Pull latest rows from disk into cache, then repaint table + stats."""
    global cached_rows
    cached_rows = load_rows()
    apply_search_filter()
    update_stats()


def select_record(event=None):
    """Load the clicked row into the form so it can be edited."""
    selection = table.selection()
    if not selection:
        return
    values = table.item(selection[0], "values")
    if not values:
        return
    reset_field_borders()
    record_id_var.set(values[0])
    name_var.set(values[1])
    email_var.set(values[2])
    country_code_var.set(values[3])
    phone_number_var.set(values[4])
    mode_label.config(text=f" EDITING #{values[0]} ", bg=PRIMARY_TINT, fg=PRIMARY_DEEP)
    set_status(f"Selected record ID {values[0]}.", is_error=False)


_hover_state = {"iid": None, "prev_tags": ()}


def _on_table_motion(event):
    """Highlight whichever row the mouse is over (skip the selected one)."""
    iid = table.identify_row(event.y)
    if iid == _hover_state["iid"]:
        return
    _clear_table_hover()
    if iid and iid not in table.selection():
        _hover_state["iid"] = iid
        _hover_state["prev_tags"] = table.item(iid, "tags")
        table.item(iid, tags=("hoverrow",))


def _clear_table_hover(event=None):
    """Restore the previously-hovered row's real tag (zebra stripe) on mouse-out."""
    iid = _hover_state["iid"]
    if iid and table.exists(iid):
        table.item(iid, tags=_hover_state["prev_tags"])
    _hover_state["iid"] = None


# ==================================================================
# CRUD actions
# ==================================================================
def add_record(event=None):
    """Validate + insert a new row, checking id/email/phone dupes first."""
    _, name, email, country_code, phone_number = get_form_values()
    if not validate_data(name, email, country_code, phone_number):
        return

    wb, sheet = get_workbook_sheet()

    record_id = get_next_id(sheet)
    if id_exists(sheet, record_id):
        # shouldn't really happen (auto id), but guard against manual edits to the sheet
        set_status(f"Record ID {record_id} already exists. Refresh and try again.", is_error=True)
        return

    if email_exists(sheet, email):
        mark_invalid(email_entry)
        set_status("A record with this email already exists.", is_error=True)
        return

    if phone_number_exists(sheet, phone_number):
        mark_invalid(phone_number_entry)
        set_status("A record with this phone number already exists.", is_error=True)
        return

    sheet.append([record_id, name, email, country_code, phone_number])
    wb.save(config.FILE_NAME)
    refresh_table()
    clear_form()
    set_status(f"Saved record ID {record_id}.", is_error=False)


def update_record():
    """Validate + overwrite the row matching the given record id."""
    record_id, name, email, country_code, phone_number = get_form_values()
    if not validate_record_id(record_id):
        return
    if not validate_data(name, email, country_code, phone_number):
        return

    record_id_int = int(record_id)
    wb, sheet = get_workbook_sheet()

    if email_exists(sheet, email, exclude_id=record_id_int):
        mark_invalid(email_entry)
        set_status("Another record already uses this email.", is_error=True)
        return

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id_int:
            row[1].value = name
            row[2].value = email
            row[3].value = country_code
            row[4].value = phone_number
            wb.save(config.FILE_NAME)
            refresh_table()
            set_status(f"Updated record ID {record_id_int}.", is_error=False)
            return

    set_status("Record ID not found.", is_error=True)


def delete_record(event=None):
    """Confirm w/ user, then delete the row matching the given record id."""
    record_id, name, _, _, _ = get_form_values()
    if not validate_record_id(record_id):
        return

    record_id_int = int(record_id)
    label = f"“{name}” (ID {record_id_int})" if name else f"record ID {record_id_int}"
    if not messagebox.askyesno(
        "Confirm Deletion",
        f"Delete {label}?\n\nThis action cannot be undone.",
        icon="warning",
        default="no",
    ):
        return

    wb, sheet = get_workbook_sheet()
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id_int:
            sheet.delete_rows(row[0].row, amount=1)
            wb.save(config.FILE_NAME)
            refresh_table()
            clear_form()
            set_status(f"Deleted record ID {record_id_int}.", is_error=False)
            return

    set_status("Record ID not found.", is_error=True)


# ==================================================================
# UI building blocks
# ==================================================================
def make_button(parent, text, command, bg, hover_bg, fg="#FFFFFF", pady=11):
    """Flat tk.Button w/ manual hover color swap (ttk styling doesn't cut it here)."""
    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=fg, activebackground=hover_bg, activeforeground=fg,
        bd=0, highlightthickness=0, relief="flat",
        padx=14, pady=pady, font=(FONT, 10, "bold"), cursor="hand2",
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=hover_bg))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


def section_title(parent, text, icon=None):
    """Heading label w/ a small accent bar on the left, optional leading icon."""
    frame = tk.Frame(parent, bg=CARD)
    tk.Frame(frame, bg=PRIMARY, width=3, height=18).pack(side="left", padx=(0, 9))
    if icon:
        tk.Label(frame, text=icon, bg=CARD, fg=PRIMARY, font=(FONT, 12)).pack(side="left", padx=(0, 6))
    tk.Label(frame, text=text, bg=CARD, fg=TEXT, font=(FONT, 13, "bold")).pack(side="left")
    return frame


def make_field(parent, grid_row, label_text, variable, required=False, hint=None, icon=None):
    """Build a labeled, icon-optional entry (+ hint text). Returns (entry, next_row)."""
    label_frame = tk.Frame(parent, bg=CARD)
    label_frame.grid(row=grid_row, column=0, sticky="w", pady=(0, 5))
    tk.Label(label_frame, text=label_text, bg=CARD, fg=TEXT_2,
             font=(FONT, 10, "bold")).pack(side="left")
    if required:
        tk.Label(label_frame, text=" *", bg=CARD, fg=DANGER,
                 font=(FONT, 10, "bold")).pack(side="left")
    grid_row += 1

    wrapper = tk.Frame(parent, bg="#FFFFFF", highlightthickness=1, highlightbackground=BORDER)
    wrapper.grid(row=grid_row, column=0, sticky="ew", pady=(0, 4 if hint else 16))
    entry_col = 0
    if icon:
        tk.Label(wrapper, text=icon, bg="#FFFFFF", fg=MUTED, font=(FONT, 10)).grid(
            row=0, column=0, padx=(10, 4), pady=8)
        entry_col = 1
    wrapper.columnconfigure(entry_col, weight=1)
    grid_row += 1

    entry = tk.Entry(
        wrapper, textvariable=variable, font=(FONT, 10),
        relief="flat", bg="#FFFFFF", fg=TEXT, insertbackground=TEXT,
        highlightthickness=0, bd=0,
    )
    entry.grid(row=0, column=entry_col, sticky="ew", padx=(10 if not icon else 0, 10), pady=8)

    # border lives on the wrapper now, so fake the same focus-glow tk.Entry gives for free
    entry.bind("<FocusIn>", lambda e: wrapper.config(highlightbackground=BORDER_FOCUS, highlightcolor=BORDER_FOCUS))
    entry.bind("<FocusOut>", lambda e: wrapper.config(highlightbackground=BORDER, highlightcolor=BORDER))
    FIELD_WRAPPERS[entry] = wrapper

    if hint:
        tk.Label(parent, text=hint, bg=CARD, fg=MUTED, font=(FONT, 8)).grid(
            row=grid_row, column=0, sticky="w", pady=(2, 16))
        grid_row += 1

    return entry, grid_row


def _apply_placeholder(event=None):
    """Show placeholder text in search box when it's empty."""
    if not search_var.get().strip():
        search_entry.config(fg=MUTED)
        search_var.set(SEARCH_PLACEHOLDER)


def _remove_placeholder(event=None):
    """Clear placeholder text as soon as the search box gets focus."""
    if search_var.get() == SEARCH_PLACEHOLDER:
        search_var.set("")
        search_entry.config(fg=TEXT)


# ==================================================================
# UI construction
# ==================================================================
def build_ui(root):
    """Wire up the whole window: styles, header, form card, list card, table."""
    root.title("Employee Contact Manager")
    root.geometry("1220x860")
    root.minsize(1060, 700)
    root.configure(bg=BG)

    style = ttk.Style()
    style.theme_use("clam")

    style.configure("TFrame", background=BG)
    style.configure("Card.TFrame", background=CARD)

    style.configure("Treeview",
                    font=(FONT, 10), rowheight=36,
                    background=CARD, fieldbackground=CARD,
                    foreground=TEXT, borderwidth=0)
    style.configure("Treeview.Heading",
                    font=(FONT, 8, "bold"),
                    background=HEAD_BG, foreground=HEAD_FG,
                    relief="flat", padding=(10, 11))
    style.map("Treeview.Heading", background=[("active", NEUTRAL_HOVER)])
    style.map("Treeview",
              background=[("selected", ROW_SELECTED)],
              foreground=[("selected", ROW_SEL_FG)])

    style.configure("Vertical.TScrollbar", background="#CBD5E1", troughcolor=CARD,
                    bordercolor=CARD, arrowsize=12)
    style.configure("Horizontal.TScrollbar", background="#CBD5E1", troughcolor=CARD,
                    bordercolor=CARD, arrowsize=12)

    # ---------------- Header ----------------
    head = tk.Frame(root, bg=HEADER_BG)
    head.pack(fill="x")
    head_inner = tk.Frame(head, bg=HEADER_BG)
    head_inner.pack(fill="x", padx=28, pady=18)

    logo_badge = tk.Label(head_inner, text="📇", bg=PRIMARY, fg="#FFFFFF",
                          font=(FONT, 16), padx=13, pady=9)
    logo_badge.pack(side="left", padx=(0, 16))

    title_col = tk.Frame(head_inner, bg=HEADER_BG)
    title_col.pack(side="left", fill="x", expand=True)
    tk.Label(title_col, text="Employee Contact Manager", bg=HEADER_BG, fg=HEADER_FG,
             font=(FONT, 18, "bold")).pack(anchor="w")
    tk.Label(title_col, text="Create, update, delete, and review employee records in one place.",
             bg=HEADER_BG, fg=HEADER_MUTED, font=(FONT, 9)).pack(anchor="w", pady=(3, 0))

    # Keyboard shortcut chips
    chips = tk.Frame(head_inner, bg=HEADER_BG)
    chips.pack(side="right", anchor="s")
    for key, action in (("Enter", "Add"), ("Esc", "Clear"), ("Del", "Delete")):
        chip = tk.Frame(chips, bg=HEADER_CHIP)
        chip.pack(side="left", padx=(8, 0))
        tk.Label(chip, text=f" {key} ", bg="#334155", fg=HEADER_FG,
                 font=(FONT, 8, "bold")).pack(side="left", padx=(4, 0), pady=4)
        tk.Label(chip, text=f"{action} ", bg=HEADER_CHIP, fg=HEADER_MUTED,
                 font=(FONT, 8)).pack(side="left", padx=(4, 6), pady=4)

    edge = tk.Frame(root, bg=BG)
    edge.pack(fill="x")
    tk.Frame(edge, bg=PRIMARY, height=2, bd=0, highlightthickness=0).pack(fill="x")
    tk.Frame(edge, bg=PRIMARY_DARK, height=1, bd=0, highlightthickness=0).pack(fill="x")

    footer = tk.Frame(root, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
    footer.pack(fill="x", side="bottom")
    tk.Label(footer, text="Employee Contact Manager  ·  records stored locally in contact.xlsx",
             bg=CARD, fg=MUTED, font=(FONT, 8)).pack(side="left", padx=28, pady=7)

    # ---------------- Main area ----------------
    main = ttk.Frame(root, style="TFrame", padding=(24, 20, 24, 24))
    main.pack(fill="both", expand=True)
    main.columnconfigure(0, weight=2, uniform="cols")
    main.columnconfigure(1, weight=3, uniform="cols")
    main.rowconfigure(0, weight=1)

    # cards sit on a "shadow" frame, offset a couple px, for a subtle lift off the bg
    form_slot = tk.Frame(main, bg=BG)
    form_slot.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
    tk.Frame(form_slot, bg="#D8DCE6").place(x=3, y=3, relwidth=1, relheight=1)
    form_outer = tk.Frame(form_slot, bg=CARD, highlightthickness=1, highlightbackground=BORDER, bd=0)
    form_outer.place(x=0, y=0, relwidth=1, relheight=1, width=-3, height=-3)
    form_card = ttk.Frame(form_outer, style="Card.TFrame", padding=(26, 22, 26, 22))
    form_card.pack(fill="both", expand=True)
    form_card.columnconfigure(0, weight=1)

    list_slot = tk.Frame(main, bg=BG)
    list_slot.grid(row=0, column=1, sticky="nsew", padx=(12, 0))
    tk.Frame(list_slot, bg="#D8DCE6").place(x=3, y=3, relwidth=1, relheight=1)
    list_outer = tk.Frame(list_slot, bg=CARD, highlightthickness=1, highlightbackground=BORDER, bd=0)
    list_outer.place(x=0, y=0, relwidth=1, relheight=1, width=-3, height=-3)
    list_card = ttk.Frame(list_outer, style="Card.TFrame", padding=(26, 22, 26, 22))
    list_card.pack(fill="both", expand=True)
    list_card.rowconfigure(2, weight=1)
    list_card.columnconfigure(0, weight=1)

    global record_id_var, name_var, email_var, country_code_var, phone_number_var
    global name_entry, email_entry, country_code_entry, phone_number_entry, phone_group
    global status_label, status_bar, table, search_var, search_entry, count_label
    global empty_label, mode_label, clear_search_btn, stats_label

    record_id_var = tk.StringVar()
    name_var = tk.StringVar()
    email_var = tk.StringVar()
    country_code_var = tk.StringVar(value="91")
    phone_number_var = tk.StringVar()

    # ---------------- Form: header ----------------
    form_head = tk.Frame(form_card, bg=CARD)
    form_head.grid(row=0, column=0, sticky="ew", pady=(0, 18))
    form_head.columnconfigure(0, weight=1)
    section_title(form_head, "Record Details", icon="🧾").grid(row=0, column=0, sticky="w")
    mode_label = tk.Label(form_head, text=" NEW RECORD ", bg=NEUTRAL_TINT, fg=TEXT_2,
                          font=(FONT, 7, "bold"), pady=3)
    mode_label.grid(row=0, column=1, sticky="e")

    row = 1
    record_id_entry, row = make_field(
        form_card, row, "Record ID", record_id_var, icon="#",
        hint="Auto-generated on Add · filled automatically when you select a row.",
    )
    name_entry, row = make_field(form_card, row, "Full Name", name_var, required=True, icon="👤")
    email_entry, row = make_field(form_card, row, "Email", email_var, required=True, icon="✉")

    # Phone group
    phone_label_frame = tk.Frame(form_card, bg=CARD)
    phone_label_frame.grid(row=row, column=0, sticky="w", pady=(0, 5))
    tk.Label(phone_label_frame, text="Phone Number", bg=CARD, fg=TEXT_2,
             font=(FONT, 10, "bold")).pack(side="left")
    tk.Label(phone_label_frame, text=" *", bg=CARD, fg=DANGER,
             font=(FONT, 10, "bold")).pack(side="left")
    row += 1

    phone_group = tk.Frame(form_card, bg="#FFFFFF", highlightthickness=1, highlightbackground=BORDER)
    phone_group.grid(row=row, column=0, sticky="ew", pady=(0, 4))
    phone_group.columnconfigure(3, weight=1)
    row += 1

    def _phone_focus_in(event):
        phone_group.config(highlightbackground=BORDER_FOCUS, highlightcolor=BORDER_FOCUS)

    def _phone_focus_out(event):
        phone_group.config(highlightbackground=BORDER, highlightcolor=BORDER)

    tk.Label(phone_group, text="+", bg="#FFFFFF", fg=MUTED,
             font=(FONT, 10, "bold")).grid(row=0, column=0, padx=(10, 0), pady=9)

    country_code_entry = tk.Entry(
        phone_group, textvariable=country_code_var, font=(FONT, 10),
        relief="flat", bg="#FFFFFF", fg=TEXT, insertbackground=TEXT,
        highlightthickness=0, bd=0, width=4, justify="center",
    )
    country_code_entry.grid(row=0, column=1, sticky="w", padx=(4, 8), pady=9)

    tk.Frame(phone_group, bg=BORDER, width=1).grid(row=0, column=2, sticky="ns", pady=7)

    phone_number_entry = tk.Entry(
        phone_group, textvariable=phone_number_var, font=(FONT, 10),
        relief="flat", bg="#FFFFFF", fg=TEXT, insertbackground=TEXT,
        highlightthickness=0, bd=0,
    )
    phone_number_entry.grid(row=0, column=3, sticky="ew", padx=(8, 10), pady=9)

    for widget in (country_code_entry, phone_number_entry):
        widget.bind("<FocusIn>", _phone_focus_in)
        widget.bind("<FocusOut>", _phone_focus_out)

    tk.Label(form_card, text="Country code and number, digits only.", bg=CARD, fg=MUTED,
             font=(FONT, 8)).grid(row=row, column=0, sticky="w", pady=(2, 20))
    row += 1

    # main CTA — only solid-color button on the form
    make_button(form_card, "＋   Add Contact", add_record, PRIMARY, PRIMARY_DARK,
                pady=12).grid(row=row, column=0, sticky="ew")
    row += 1

    # secondary actions
    button_row = tk.Frame(form_card, bg=CARD)
    button_row.grid(row=row, column=0, sticky="ew", pady=(8, 0))
    button_row.columnconfigure((0, 1, 2), weight=1)
    row += 1

    make_button(button_row, "✎  Update", update_record,
                SUCCESS_TINT, SUCCESS_HOVER, fg=SUCCESS).grid(row=0, column=0, sticky="ew", padx=(0, 6))
    make_button(button_row, "🗑  Delete", delete_record,
                DANGER_TINT, DANGER_HOVER, fg=DANGER).grid(row=0, column=1, sticky="ew", padx=6)
    make_button(button_row, "↺  Clear", clear_form,
                NEUTRAL_TINT, NEUTRAL_HOVER, fg=NEUTRAL_TEXT).grid(row=0, column=2, sticky="ew", padx=(6, 0))

    # status pill w/ colored left edge
    status_frame = tk.Frame(form_card, bg=SUCCESS_TINT)
    status_frame.grid(row=row, column=0, sticky="ew", pady=(22, 0))
    status_frame.columnconfigure(1, weight=1)
    status_bar = tk.Frame(status_frame, bg=SUCCESS, width=3)
    status_bar.grid(row=0, column=0, sticky="ns")
    status_label = tk.Label(status_frame, text="  ✓  Ready.", bg=SUCCESS_TINT, fg=SUCCESS,
                            font=(FONT, 10, "bold"), anchor="w", pady=10)
    status_label.grid(row=0, column=1, sticky="ew")

    # ---------------- List: header + search ----------------
    list_head = tk.Frame(list_card, bg=CARD)
    list_head.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    list_head.columnconfigure(0, weight=1)

    title_col2 = tk.Frame(list_head, bg=CARD)
    title_col2.grid(row=0, column=0, sticky="w")
    title_line = tk.Frame(title_col2, bg=CARD)
    title_line.pack(anchor="w")
    section_title(title_line, "Saved Records", icon="📋").pack(side="left")
    count_label = tk.Label(title_line, text="  0 contacts  ", bg=PRIMARY_TINT, fg=PRIMARY_DEEP,
                           font=(FONT, 8, "bold"), pady=3)
    count_label.pack(side="left", padx=(10, 0))
    stats_label = tk.Label(title_col2, text="", bg=CARD, fg=MUTED, font=(FONT, 8))
    stats_label.pack(anchor="w", pady=(4, 0))

    search_var = tk.StringVar()
    search_box = tk.Frame(list_head, bg="#FFFFFF", highlightthickness=1, highlightbackground=BORDER)
    search_box.grid(row=0, column=1, sticky="e")
    search_box.columnconfigure(1, weight=1)

    tk.Label(search_box, text="🔍", bg="#FFFFFF", fg=MUTED,
             font=(FONT, 10)).grid(row=0, column=0, padx=(10, 2), pady=7)

    search_entry = tk.Entry(
        search_box, textvariable=search_var, font=(FONT, 10),
        relief="flat", bg="#FFFFFF", fg=TEXT, insertbackground=TEXT,
        highlightthickness=0, width=28,
    )
    search_entry.grid(row=0, column=1, padx=(0, 2), pady=7)
    search_entry.bind("<KeyRelease>", apply_search_filter)
    search_entry.bind("<FocusIn>", lambda e: (_remove_placeholder(),
                      search_box.config(highlightbackground=BORDER_FOCUS)))
    search_entry.bind("<FocusOut>", lambda e: (_apply_placeholder(),
                      search_box.config(highlightbackground=BORDER)))

    clear_search_btn = tk.Label(search_box, text="✕", bg="#FFFFFF", fg=MUTED,
                                font=(FONT, 9, "bold"), cursor="hand2")
    clear_search_btn.grid(row=0, column=2, padx=(0, 10), pady=7)
    clear_search_btn.bind("<Button-1>", lambda e: clear_search())
    clear_search_btn.grid_remove()

    tk.Frame(list_card, bg=BORDER, height=1).grid(row=1, column=0, sticky="ew", pady=(14, 14))

    # ---------------- Table ----------------
    table_frame = tk.Frame(list_card, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
    table_frame.grid(row=2, column=0, sticky="nsew")
    table_frame.rowconfigure(0, weight=1)
    table_frame.columnconfigure(0, weight=1)

    table = ttk.Treeview(table_frame, columns=COLUMNS, show="headings", selectmode="browse")
    for col in COLUMNS:
        table.heading(col, text=HEADINGS[col], command=lambda c=col: sort_by(c))
    table.column("id", width=60, anchor="center", stretch=False)
    table.column("name", width=170, anchor="w")
    table.column("email", width=230, anchor="w")
    table.column("country_code", width=70, anchor="center", stretch=False)
    table.column("phone_number", width=140, anchor="center", stretch=False)
    table.grid(row=0, column=0, sticky="nsew")
    table.bind("<<TreeviewSelect>>", select_record)
    table.bind("<Delete>", delete_record)
    table.bind("<Motion>", _on_table_motion)
    table.bind("<Leave>", _clear_table_hover)
    table.tag_configure("oddrow", background=ROW_ODD)
    table.tag_configure("evenrow", background=ROW_EVEN)
    table.tag_configure("hoverrow", background=NEUTRAL_HOVER)

    empty_label = tk.Label(table_frame, text="No contacts yet.", bg=CARD, fg=MUTED,
                           font=(FONT, 11), justify="center")

    scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
    scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
    table.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
    scrollbar_y.grid(row=0, column=1, sticky="ns")
    scrollbar_x.grid(row=1, column=0, sticky="ew")

    tk.Label(list_card, text="Click a column header to sort · click a row to load it into the form.",
             bg=CARD, fg=MUTED, font=(FONT, 8)).grid(row=3, column=0, sticky="w", pady=(10, 0))

    # ---------------- Keyboard shortcuts ----------------
    for widget in (name_entry, email_entry, country_code_entry, phone_number_entry):
        widget.bind("<Return>", add_record)
    root.bind("<Escape>", clear_form)

    _apply_placeholder()
    refresh_table()
    clear_form()
    set_status(f"Loaded {len(cached_rows)} record(s).", is_error=False)


if __name__ == "__main__":
    root = tk.Tk()
    build_ui(root)
    root.mainloop()